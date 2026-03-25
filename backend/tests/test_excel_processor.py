from io import BytesIO

import openpyxl


def _wb_to_bytes(wb: openpyxl.Workbook) -> bytes:
    bio = BytesIO()
    wb.save(bio)
    return bio.getvalue()


def test_excel_processor__multi_sheet_leading_zeros():
    from app.excel_processor import process_excel_file

    wb = openpyxl.Workbook()

    # Sheet 1: has a title row, then header row with "Serial Number"
    ws1 = wb.active
    ws1.title = "Alcon"
    ws1.append(["Used_lens"])  # title row (non-header)
    ws1.append(["Serial Number", "Other"])
    ws1.append(["00123", "x"])
    ws1.append(["", "y"])  # empty serial should be skipped
    ws1.append(["04567", "z"])

    # Sheet 2: header uses "SN"
    ws2 = wb.create_sheet("AMO")
    ws2.append(["SN", "foo"])
    ws2.append(["90001", "a"])

    out = process_excel_file(_wb_to_bytes(wb))
    assert out == [
        {"serial_number": "00123", "sheet_name": "Alcon"},
        {"serial_number": "04567", "sheet_name": "Alcon"},
        {"serial_number": "90001", "sheet_name": "AMO"},
    ]


def test_excel_processor__fallback_first_nonempty_column():
    from app.excel_processor import process_excel_file

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "BrandX"
    ws.append(["foo", "bar"])
    ws.append(["111", "x"])
    ws.append(["222", None])

    out = process_excel_file(_wb_to_bytes(wb))
    assert out == [
        {"serial_number": "111", "sheet_name": "BrandX"},
        {"serial_number": "222", "sheet_name": "BrandX"},
    ]


def test_excel_processor__invalid_bytes_raises():
    from app.excel_processor import process_excel_file

    try:
        process_excel_file(b"not an excel file")
        assert False, "expected exception"
    except Exception as e:
        assert "Failed to read Excel file" in str(e)


def test_excel_processor__skips_empty_sheets_and_uses_code_column_fallback():
    from app.excel_processor import process_excel_file

    wb = openpyxl.Workbook()
    ws1 = wb.active
    ws1.title = "EmptySheet"
    # header only -> empty after read
    ws1.append(["Serial Number"])

    ws2 = wb.create_sheet("CodeSheet")
    ws2.append(["product code", "x"])
    ws2.append(["ABC123", "1"])
    ws2.append(["DEF456", "2"])

    out = process_excel_file(_wb_to_bytes(wb))
    assert out == [
        {"serial_number": "ABC123", "sheet_name": "CodeSheet"},
        {"serial_number": "DEF456", "sheet_name": "CodeSheet"},
    ]


def test_excel_processor__sheet_with_no_usable_columns_is_skipped():
    from app.excel_processor import process_excel_file

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Blank"
    ws.append(["a", "b"])
    # rows with Nones -> no non-null values
    ws.append([None, None])
    ws.append([None, None])

    out = process_excel_file(_wb_to_bytes(wb))
    assert out == []


