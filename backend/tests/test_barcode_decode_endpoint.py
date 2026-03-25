from __future__ import annotations

from io import BytesIO
from pathlib import Path
from typing import Iterable

import pytest


TESTSN_XLSX = Path(__file__).with_name("testsn.xlsx")

if not TESTSN_XLSX.exists():
    pytest.skip("testsn.xlsx not found; skipping /barcode/decode image tests", allow_module_level=True)


def _load_testsn_barcodes(xlsx_path: Path) -> list[str]:
    """
    Load barcode ground-truth strings from testsn.xlsx.

    Expected header contains a 'barcode' column.
    """
    try:
        from openpyxl import load_workbook
    except Exception as e:  # pragma: no cover
        pytest.skip(f"openpyxl not available: {e}", allow_module_level=True)

    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    headers = [(str(v).strip().lower() if v is not None else "") for v in header_row]
    if "barcode" not in headers:
        raise AssertionError(f"Expected a 'barcode' column in {xlsx_path}, got headers={headers}")

    barcode_idx = headers.index("barcode")
    out: list[str] = []

    def _cell_to_str(v) -> str:
        if v is None:
            return ""
        # Preserve integers without ".0"
        if isinstance(v, int):
            return str(v)
        if isinstance(v, float) and v.is_integer():
            return str(int(v))
        return str(v)

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or barcode_idx >= len(row):
            continue
        raw = _cell_to_str(row[barcode_idx]).strip()
        if not raw:
            continue
        out.append(raw)

    if not out:
        raise AssertionError(f"No barcode values found in {xlsx_path}")
    return out


def _render_code128_png(data: str) -> bytes:
    """
    Render a Code128 barcode image as PNG bytes.

    We use Code128 because it supports a wide range of ASCII and is supported by ZBar/pyzbar.
    """
    try:
        from barcode import get_barcode_class
        from barcode.writer import ImageWriter
    except Exception as e:  # pragma: no cover
        pytest.skip(f"python-barcode not available: {e}")

    code128 = get_barcode_class("code128")
    b = code128(data, writer=ImageWriter())

    buf = BytesIO()
    # Settings chosen to be easy for decoders (adequate quiet zone + size, no text).
    b.write(
        buf,
        options={
            "write_text": False,
            "quiet_zone": 6.5,
            "module_width": 0.25,
            "module_height": 20.0,
            "dpi": 300,
        },
    )
    return buf.getvalue()


@pytest.fixture(scope="module")
def client():
    """
    FastAPI test client for calling the decode endpoint without running uvicorn.
    """
    from fastapi.testclient import TestClient

    # Import app.main to access the FastAPI app and engine availability flags.
    import app.main as main_mod

    # The endpoint can run without pyzbar (OpenCV/other engines), but Code128 decoding is most
    # reliably exercised via pyzbar. Skip on machines without it to avoid flaky results.
    if not getattr(main_mod, "_PYZBAR_AVAILABLE", False):
        pytest.skip("pyzbar/ZBar not available; skipping /barcode/decode image tests")

    return TestClient(main_mod.app)


@pytest.mark.parametrize("barcode_value", _load_testsn_barcodes(TESTSN_XLSX), ids=lambda s: s[:30])
def test_barcode_decode__images_from_testsn_xlsx_roundtrip(client, barcode_value: str):
    png = _render_code128_png(barcode_value)

    resp = client.post(
        "/barcode/decode",
        files={"image": ("barcode.png", png, "image/png")},
    )
    assert resp.status_code == 200, resp.text
    payload = resp.json()

    assert isinstance(payload, dict)
    assert payload.get("count", 0) >= 1
    results = payload.get("results") or []
    texts = [r.get("text") for r in results if isinstance(r, dict)]

    assert barcode_value in texts


